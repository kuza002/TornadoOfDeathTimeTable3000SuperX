from openpyxl.styles import PatternFill

from parse_time_table import Parser
import xlsxwriter
import shutil
import pickle
import openpyxl
import numpy as np

days_of_week = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб']
times = ['8.30-10.00', '10.10-11.40', '12.10-13.40', '13.50-15.20', '15.50-17.20', '17.30-19.00', '19.10-20.40']
our_classrooms = ['802', '804', '808', '809', '810', '811', '910', '1009', '1111', '1112', '1206']
classrooms_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
good_color = 'ccffcc'
bad_color = 'ffadd6'


def paint_cells(sheet, lessons, color, magic_var, column):
    for lesson in lessons:

        dw_index = days_of_week.index(lesson.day_of_week)

        time_index = times.index(lesson.duration)

        lesson_row = magic_var[dw_index] + time_index * 2

        if 'н/н' in lesson.cell.value:
            sheet[column + str(lesson_row)].fill = PatternFill("solid", start_color=color)
        elif 'ч/н' in lesson.cell.value:
            sheet[column + str(lesson_row + 1)].fill = PatternFill("solid", start_color=color)
        else:
            sheet[column + str(lesson_row)].fill = PatternFill("solid", start_color=color)
            sheet[column + str(lesson_row + 1)].fill = PatternFill("solid", start_color=color)


# Get data from timetable
data = Parser()

with open('filename.pickle', 'wb') as output:
    pickle.dump(data, output)

# with open('filename.pickle', 'rb') as handle:
#     data = pickle.load(handle)

# groups = input("Введите нужные группы через запятую с пробелом, точно как в расписании: ") \
#     .strip().lower().split(", ")

groups = ['09-933 (1)', '09-012 (1)', '09-022 (2)', '09-145', '09-125 (2)']

# Getting lessons for all groups

all_lessons = {}

for group in groups:
    lessons_for_group = data.get_lessons_by_group(group)
    all_lessons[group] = lessons_for_group

# Copy template


out_file = 'output.xlsx'
shutil.copyfile('template.xlsx', out_file)

# Create area to put lessons ( sry about this ;) )

magic_var = np.array([i for i in range(6, 85, 14)])

group_coord = 'C'

wb = openpyxl.load_workbook(out_file)
sheet = wb['Расписание лаборантов 21-22 1се']

# Create table for workers
print("Start generate table for workers")

for group, lessons in all_lessons.items():
    sheet[group_coord + "2"] = group
    for lesson in lessons:
        dw_index = days_of_week.index(lesson.day_of_week)

        time_index = times.index(lesson.duration)

        lesson_row = magic_var[dw_index] + time_index * 2

        if 'н/н' in lesson.cell.value:
            sheet[group_coord + str(lesson_row)].fill = PatternFill("solid", start_color=bad_color)
        elif 'ч/н' in lesson.cell.value:
            sheet[group_coord + str(lesson_row + 1)].fill = PatternFill("solid", start_color=bad_color)
        else:
            sheet[group_coord + str(lesson_row)].fill = PatternFill("solid", start_color=bad_color)
            sheet[group_coord + str(lesson_row + 1)].fill = PatternFill("solid", start_color=bad_color)

    group_coord = chr(ord(group_coord) + 1)

print("Table for worker is done")
print("Start create table by classroom")

all_lessons_by_classroom = data.get_lessons_by_classrooms(our_classrooms)

magic_var = np.array([i for i in range(4, 76, 14)])

for classroom, lessons in all_lessons_by_classroom.items():
    classroom_index = our_classrooms.index(classroom)
    column = classrooms_columns[classroom_index]
    paint_cells(wb[wb.sheetnames[0]], lessons, good_color, magic_var, column)

wb.save(out_file)
