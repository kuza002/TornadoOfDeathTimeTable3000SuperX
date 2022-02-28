import datetime
import os
import shutil
import threading
import tkinter as tk
import openpyxl
from tkinter import BOTH, filedialog, PhotoImage, Menu, messagebox as mbox
from tkinter.ttk import Frame, Button, Style, Entry, Label
from Settings_window import Worker_set
from parse_time_table import Parser
from support_file import *


def thread(fn):
    def execute(*args, **kwargs):
        threading.Thread(target=fn, args=args, kwargs=kwargs).start()
    return execute


class Example(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)
        self.data = None
        self.parent = parent
        self.parsed = False
        self.wb = openpyxl.load_workbook('some_files/template.xlsx')
        self.__initUI__()

    def __initUI__(self):
        self.parent.title("Шото с чем-то")
        self.parent.iconphoto(False, PhotoImage(file='some_files/icon.png'))
        self.style = Style()
        self.style.theme_use("default")
        self.pack(fill=BOTH, expand=1)
        self.menubar = Menu(self)
        self.menubar.add_cascade(label="Сотрудники", command=self.__open_worker_settings__)
        self.parent.config(menu=self.menubar)
        bias = 100
        Label(self, text='Расположение файла с расписанием занятий').place(x=20 + bias, y=70)
        self.file_path = Entry(self)
        self.file_path.place(x=20 + bias, y=90, width=267, height=30)

        self.file_picker = Button(self, text="Выбрать файл", command=self.__open_file__)
        self.file_picker.place(x=20 + bias, y=130)

        self.parse_button = Button(self, text="Спарсить", command=self.__parse_file__)
        self.parse_button.place(x=200 + bias, y=130)

        self.just_button = Button(self, text='Работать!', command=self.__make_table__)
        self.just_button.place(x=120 + bias, y=450)

    def __open_file__(self):
        ftypes = [('Excel files', '*.xlsx')]
        dlg = filedialog.Open(self, filetypes=ftypes)

        fl = dlg.show()
        self.file_path.insert(0, fl)

    def __open_worker_settings__(self):
        window = Worker_set(self)
        window.title('Сотрудники')
        window.geometry("300x300")
        window.grab_set()
        window.iconphoto(False, PhotoImage(file='some_files/icon.png'))

    @thread
    def __parse_file__(self):
        self.parse_button.config(state=tk.DISABLED)
        self.just_button.config(state=tk.DISABLED)
        self.file_picker.config(state=tk.DISABLED)
        if os.path.exists('some_files/data.pickle'):
            shutil.copyfile('some_files/data.pickle', 'some_files/old_data.pickle')
        if self.file_path.get() == '':
            mbox.showerror("Ошибка!", 'Выберите сначала файл!')
        else:
            self.data = Parser(self.file_path.get())
            with open('some_files/data.pickle', 'wb') as output:
                pickle.dump(self.data, output)
            self.parsed = True
        self.parse_button.config(state=tk.NORMAL)
        self.just_button.config(state=tk.NORMAL)
        self.file_picker.config(state=tk.NORMAL)

        mbox.showinfo(message='Парсинг закончен!')

    def __make_table__(self):
        if not self.parsed:
            if os.path.exists('some_files/data.pickle'):
                with open('some_files/data.pickle', 'rb') as f:
                    self.data = pickle.load(f)
            else:
                mbox.showerror('Ошибка!', 'Не найден файл data.pickle')
        wb = openpyxl.load_workbook('some_files/template.xlsx')

        with open('some_files/workers.pickle', 'rb') as f:
            workers = pickle.load(f)
        # Create area to put lessons ( sry about this ;) )
        magic_var = list(range(6, 77, 14))

        # Getting lessons for all groups
        all_lessons = {}

        for worker in workers:
            if worker[1]=='':
                mbox.showwarning(title='Ошибка',
                                 message=f'У сотрудника {worker[0]} нет группы.\n'
                                         f'Он не будет сохранен в таблице')

            elif self.data.get_lessons_by_group(worker[1])==None:
                mbox.showerror(title='Ошибка!',
                               message=f'Не удалось найти группу сотрудника {worker[0]}\n'
                                       f'Он не будет сохранен в таблице')
            else:
                all_lessons[':'.join(worker)] = self.data.get_lessons_by_group(worker[1])

        # region Create table for workers
        group_coord = 'C'
        for worker, lessons in all_lessons.items():
            worker = worker.split(':')
            wb[wb.sheetnames[1]][group_coord + "1"] = worker[0]
            wb[wb.sheetnames[1]][group_coord + "2"] = worker[1]
            paint_cells(wb[wb.sheetnames[1]], lessons, bad_color, magic_var, group_coord, True)

            group_coord = chr(ord(group_coord) + 1)
        # endregion

        # region Create table for classrooms
        all_lessons_by_classroom = self.data.get_lessons_by_classrooms(our_classrooms)

        for classroom, lessons in all_lessons_by_classroom.items():
            classroom_index = our_classrooms.index(classroom)
            column = classrooms_columns[classroom_index]
            paint_cells(wb[wb.sheetnames[0]], lessons, good_color, magic_var, column)
        # endregion

        directory = filedialog.askdirectory()
        if directory:
            now = datetime.datetime.now()
            date = f'{now.day}.{now.month}.{now.year}'
            wb.save(f'{directory}/schedule_{date}.xlsx')
            ans=mbox.askquestion(title='Сохранение', message='Таблица сохранена!\nЗакрыть приложение?')

            if ans=='yes':
                self.quit()
        else:
            mbox.showwarning(title='Сохранение', message='Файл не сохранён!')
        print('DONE!!!')
