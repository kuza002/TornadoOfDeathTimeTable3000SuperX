import pickle

import numpy as np
from openpyxl.styles import PatternFill

days_of_week = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб']
times = ['8.30-10.00', '10.10-11.40', '11.50-13.20', '14.00-15.30', '15.40-17.10', '17.50-19.20']
our_classrooms = ['802', '804', '808', '809', '810', '811', '910', '1009', '1111', '1112', '1206']
classrooms_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
good_color = 'ccffcc'
bad_color = 'ffadd6'


def write_in_file(path, var):
    with open(path, 'a') as file:
        file.write(var)


def paint_cells(sheet, lessons, color, magic_var, column, flag_for_groups=False):
    for lesson in lessons:
        try:
            dw_index = days_of_week.index(lesson.day_of_week)

            if flag_for_groups:
                if lesson.day_of_week == 'сб' and times.index(lesson.duration) > 3:
                    continue
            time_index = times.index(lesson.duration)

            lesson_row = magic_var[dw_index] + time_index * 2

            if 'н/н' in lesson.cell.value:
                sheet[column + str(lesson_row)].fill = PatternFill("solid", start_color=color)
            elif 'ч/н' in lesson.cell.value:
                sheet[column + str(lesson_row + 1)].fill = PatternFill("solid", start_color=color)
            else:
                sheet[column + str(lesson_row)].fill = PatternFill("solid", start_color=color)
                sheet[column + str(lesson_row + 1)].fill = PatternFill("solid", start_color=color)
        except:
            write_in_file('some_files/log.txt', f'{lesson.cell.coordinate, lesson.cell.value} is a lesson after 19:30')
            continue


class Lesson:
    def __init__(self, cell, duration, day_of_week, group_number):
        self.cell = cell
        self.duration = duration
        self.day_of_week = day_of_week
        self.group_number = group_number


if __name__ == '__main__':
    groups = [
        ['Илья Г.', '09-933 ', '(1)'],
        ['Данил',   '09-933 ', '(1)'],
        ['Ахад',    '09-063 ', '(1)'],
        ['Максим',  '09-012 ', '(1)'],
        ['Илья К.', '09-012 ', '(1)'],
    ]
    with open('some_files/workers.pickle', 'wb') as f:
        pickle.dump(groups, f)
